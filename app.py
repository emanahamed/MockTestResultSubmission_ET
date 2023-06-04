from datetime import datetime
import os
from flask import Flask, render_template, request, redirect, url_for
from flask_mail import Mail, Message
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'onemorezero00@gmail.com'  # Replace with your Gmail email address
app.config['MAIL_PASSWORD'] = 'cihdwimjdatnhegu'  # Replace with your Gmail password or app password

mail = Mail(app)


def create_excel_file(data):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header row
    header = [
        'ID', 'Name of Student', 'Year Group', 'Day', 'Subject', 'Mark Scored',
        'Total Mark', 'Score as %', 'Tier', 'Grade', 'Full Paper'
    ]
    for col_num, header_value in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header_value

    # Write the data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_value

        # # Apply conditional formatting
        # fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        # fill_golden = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        #
        # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):
        #     for cell in row:
        #         percentage = int(cell.value)  # Convert to integer
        #         if 85 <= percentage <= 89:
        #             cell.fill = fill_golden
        #         elif percentage > 90:
        #             cell.fill = fill_green

    # Save the workbook to a temporary file
    filename = 'data.xlsx'
    workbook.save(filename)
    return filename


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        data = []
        tutors_name = request.form.get('tutors_name')
        tutors_email = request.form.get('tutors_email')
        department = request.form.get('department')

        for i in range(len(request.form.getlist('id[]'))):
            row = [
                request.form.getlist('id[]')[i],
                request.form.getlist('name[]')[i],
                request.form.getlist('year[]')[i],
                request.form.getlist('day[]')[i],
                request.form.getlist('subject[]')[i],
                request.form.getlist('mark[]')[i],
                request.form.getlist('total[]')[i],
                request.form.getlist('score[]')[i],
                request.form.getlist('tier[]')[i],
                request.form.getlist('grade[]')[i],
                request.form.getlist('paper[]')[i]
            ]
            data.append(row)

        # Create Excel file
        filename = create_excel_file(data)

        # Send email
        recipient = 'management@exceltutors.org.uk'
        subject = f'{tutors_name}_{department} Mock Result Data Submission'
        body = f'This is the mock test result submitted on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} by {tutors_name} from {department} department .'
        msg = Message(subject=subject, sender=app.config['MAIL_USERNAME'], recipients=[recipient])
        msg.body = body
        with app.open_resource(filename) as attachment:
            msg.attach(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', attachment.read())
        mail.send(msg)

        # Remove the temporary file
        os.remove(filename)

        # Send confirmation email to the sender
        sender_email = request.form.get('tutors_email')
        confirmation_subject = 'Mock Result Submission Confirmation'
        confirmation_body = 'Thank you for submitting the student data. Your submission has been received and will be processed.'
        confirmation_msg = Message(subject=confirmation_subject, sender=app.config['MAIL_USERNAME'], recipients=[sender_email])
        confirmation_msg.body = confirmation_body
        mail.send(confirmation_msg)

        return redirect(url_for('index'))

    return render_template('index.html')


if __name__ == '__main__':
    app.run(debug=True)
